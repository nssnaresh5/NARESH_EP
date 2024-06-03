import pandas as pd
from openpyxl import load_workbook


class ExcelManipulation:
    def __init__(self):
        self.input_data = None
        self.input_excel_path = "ESXP_Param_v1.xlsx"
        self.device_list_excel_path = "Device List ESXP-2.36.xlsx"

    def read_input_file(self):
        book = load_workbook(self.input_excel_path, read_only=True)
        with pd.ExcelFile(book, engine='openpyxl') as xls:
            df = pd.read_excel(xls, sheet_name='Param', header=[0])
        data_dict = df.to_dict(orient='list')
        self.input_data = []
        num_rows = len(next(iter(data_dict.values())))
        for i in range(num_rows):
            new_dict = {}
            for key, values in data_dict.items():
                if key not in new_dict:
                    new_dict[key] = values[i]
            self.input_data.append(new_dict)
        return self.input_data

    # def read_device_list(self, cr):
    #     book = load_workbook(self.device_list_excel_path, read_only=True)
    #     with pd.ExcelFile(book, engine='openpyxl') as xls:
    #         df = pd.read_excel(xls, sheet_name='DEVICE LIST', header=[1])
    #     data_dict = df.to_dict(orient='list')
    #     self.input_data = []
    #     num_rows = len(next(iter(data_dict.values())))
    #     for i in range(num_rows):
    #         new_dict = {}
    #         for key, values in data_dict.items():
    #             if key not in new_dict:
    #                 new_dict[key] = values[i]
    #         self.input_data.append(new_dict)
    #     for c in self.input_data:
    #         if c['COMMERCIAL\nREFERENCE'] == cr:
    #             return c
    #
    #     return self.input_data
    def read_device_list(self, cr):
        book = load_workbook(self.device_list_excel_path, read_only=True)
        with pd.ExcelFile(book, engine='openpyxl') as xls:
            df = pd.read_excel(xls, sheet_name='DEVICE LIST', header=[1])
        data_list = df.values.tolist()
        self.input_data = []
        for row in data_list:
            self.input_data.append(row)
        for c in self.input_data:
            if c[9] == cr:  # assuming the 'COMMERCIAL\nREFERENCE' is the first column
                return c
        return self.input_data
